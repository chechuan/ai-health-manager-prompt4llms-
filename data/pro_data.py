import openpyxl as op
import sys
from pathlib import Path
sys.path.append(str(Path(__file__).parent.parent.absolute()))
from jiahe_util import callEmbedding

def read_dish_xlsx(file_path="/Users/yuanhuachao/Desktop/ai-health-manager-prompt4llms/data/dishes.xlsx"):
    workbook = op.load_workbook(file_path)
    worksheet = workbook["Sheet1"]
    rows = worksheet.max_row
    columns = worksheet.max_column
    dishes = []
    for i in range(1, rows+1):
        if worksheet.cell(row=i, column=1).value.strip() == 'code':
            continue
        dishes.append(worksheet.cell(row=i, column=2).value.strip())
    import pdb
    pdb.set_trace()
    callEmbedding(dishes)

read_dish_xlsx()