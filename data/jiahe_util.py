from src.utils.api_protocal import *
from data.constrant import *
from src.prompt.model_init import callLLM, acallLLM
from src.utils.Logger import logger
import json, time
import openai
import openpyxl as op
# from sklearn.metrics.pairwise import cosine_similarity
from data.jiahe_prompt import *
from data.jiahe_constant import *

jiahe_userInfo_map = {
    'age':'年龄',
    'gender':'性别',
    'height':'身高',
    'weight':'体重',
    'manage_object':'管理目标',
    'disease':'现患疾病',
    'special_diet':'特殊口味偏好',
    'allergy_food':'过敏食物',
    'taste_preference':'口味偏好',
    'is_specific_menstrual_period':'是否特殊生理期',
    'constitution': '中医体质',
    'taste_taboo': '口味禁忌'
}


def get_userInfo_history(userInfo, history=[]):
    user_info = JiaheUserProfile().model_dump()
    for i in userInfo:
        if userInfo[i]:
            user_info[i] = userInfo[i]
    history = [
        {"role": role_map.get(str(i["role"]), "user"), "content": i["content"]}
        for i in history[-5:]     # 保留最后5轮用户信息
    ]
    his_prompt = "\n".join(
        [
            ("assistant" if not i["role"] == "user" else "user")
            + f": {i['content']}"
            for i in history
        ]
    )
    info = ''
    for i in user_info.keys():
        if user_info[i] and user_info[i] != '未知':
            info += f'{jiahe_userInfo_map[i]}：{user_info[i]}\n'

    return info, his_prompt


def get_userInfo(userInfo):
    user_info = JiaheUserProfile().model_dump()
    for i in userInfo:
        if userInfo[i]:
            user_info[i] = userInfo[i]
    info = ''
    for i in user_info.keys():
        if user_info[i] and user_info[i] != '未知':
            info += f'{jiahe_userInfo_map[i]}：{user_info[i]}\n'

    return info


def get_familyInfo_history(familyInfo, history):
    infos = ''
    roles = ''
    for user in familyInfo:
        user_info = JiaheUserProfile().model_dump()
        userInfo = user.get('userInfo', {})
        for i in userInfo:
            if userInfo[i]:
                user_info[i] = userInfo[i]
        info = f'{user.get("family_role", "")}的健康标签'
        roles = roles + user.get("family_role", "") + '，'
        for i in user_info.keys():
            if user_info[i] and user_info[i] != '未知':
                info += f'{jiahe_userInfo_map[i]}：{user_info[i]}\n'
        infos = infos + '\n' + info
    roles = roles[:-1] if roles[-1] == '，' else roles
    history = [
        {"role": role_map.get(str(i["role"]), "user"), "content": i["content"]}
        for i in history[-5:]       # 保留最后5轮用户信息
    ]
    his_prompt = "\n".join(
        [
            ("assistant" if not i["role"] == "user" else "user")
            + f": {i['content']}"
            for i in history
        ]
    )

    return roles, infos, his_prompt

def callEmbedding(
    inputs, model='bce-embedding-base-v1'
):

    client = openai.OpenAI()
    if not isinstance(inputs, list):
        inputs = [inputs]
    t_st = time.time()
    logger.debug('begin to call embedding')
    completion = client.embeddings.create(input = inputs, model=model)
    logger.debug('finished call embedding')
    time_cost = round(time.time() - t_st, 1)
    # f = open('dish_embedding', 'w')
    xs = []
    for emb in completion.data:
        xs.append(emb.embedding)
        # file.write(json.dumps(emb.embedding) + '\n')
        # logger.debug(f"get embedding {i}")

    logger.info(
        f"cost: {time_cost}s"
    )
    return xs

# def read_dish_xlsx(file_path="/Users/yuanhuachao/Desktop/ai-health-manager-prompt4llms/data/dishes.xlsx"):
#     workbook = op.load_workbook(file_path)
#     worksheet = workbook["Sheet1"]
#     rows = worksheet.max_row
#     columns = worksheet.max_column
#     dishes = []
#     with open('2.txt', 'w') as f:
#         i = 0
#         for i in range(1, rows+1):
#             # if i > 3:
#             #     break
#             if worksheet.cell(row=i, column=1).value.strip() == 'code':
#                 continue
#             xs = callEmbedding(i, worksheet.cell(row=i, column=2).value.strip(), f)
#             for x in xs:
#                 # import pdb
#                 # pdb.set_trace()
#                 f.write(json.dumps(x) + '\n')
#             # dishes.append(worksheet.cell(row=i, column=2).value.strip())
#         # f.close()
#         print(f'len dishes: {len(dishes)}')
#     return {'result':'OK'}


def get_dish_info(file):
    data = json.load(open(file, 'r'))
    return data


def get_em(f):
    d = open(f, 'r').readlines()
    print(f'emb数据条数：{len(d)}')

    def get_dish_from_database(dish):
        # get_em('dish_embedding')
        # return
        # return read_dish_xlsx()
        # # # inputs =
        logger.debug(
            "bce embedding模型输入： " + json.dumps(dish, ensure_ascii=False)
        )

        d = set(dish)
        ds = get_dish_info("dishes.json")
        idxes = []
        for i, x in enumerate(ds):
            if len(set(d) & set(x['name'])) > 0:
                idxes.append(i)

        embs = open('emb', 'r').readlines()
        for i in idxes:
            emb = json.loads(embs[i].strip())

def jaccard_text_similarity(text1, text2):
    set1 = set(text1.split())
    set2 = set(text2.split())
    return len(set1 & set2) / len(set1 | set2)

def get_dish_from_database(dish, userInfo):
    # get_em('dish_embedding')
    # return
    # return read_dish_xlsx()
    # # # inputs =

    # 1. 向量/关键词匹配
    # logger.debug(
    #     "bce embedding模型输入： " + json.dumps(dish, ensure_ascii=False)
    # )

    ds = get_dish_info(DISH_FILE)
    target_dish = {}
    # import pdb
    # pdb.set_trace()
    for i, x in enumerate(ds):
        if jaccard_text_similarity(x['name'], dish) < 0.5:
            continue
        if not target_dish and jaccard_text_similarity(x['name'], dish) >= 0.5:
            target_dish = x
        elif jaccard_text_similarity(x['name'], dish) > jaccard_text_similarity(x['name'], target_dish['name']):
            target_dish = x
    logger.debug(f'dish dataset recall data: {json.dumps(target_dish)}')
    # embs = open('emb', 'r').readlines()
    # for i in idxes:
    #     cur_emb = callEmbedding(dish)[0]
    #     emb = json.loads(embs[i].strip())
    #     # cal cos_similarity
    #     cosine_sim = cosine_similarity([cur_emb], [emb])
    #     if cosine_sim > 0.8:
    #         res.append(i)
    if not target_dish:
        return {}
    # 2. llm判断

    userInfo = get_userInfo(userInfo)
    messages = [
        {
            "role": "user",
            "content": jiahe_judge_dishes_prompt.format(
                userInfo, target_dish['name']
            ),
        }
    ]
    logger.debug(
        "儿童菜品llm判断模型输入： " + json.dumps(messages, ensure_ascii=False)
    )
    start_time = time.time()
    generate_text = callLLM(
        history=messages,
        max_tokens=512,
        top_p=0.9,
        temperature=0.8,
        do_sample=True,
        # stream=True,
        model="Qwen1.5-32B-Chat",
    )
    logger.debug("儿童菜品llm判断模型输出latancy： " + str(time.time() - start_time))
    logger.debug("儿童菜品llm判断模型输出： " + generate_text)
    if '否' in generate_text:
        return {}
    else:
        return target_dish



